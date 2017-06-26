#!/usr/local/bin/python3

from openpyxl import load_workbook
import argparse
import subprocess
import datetime
import time

parser = argparse.ArgumentParser(description= 'Takes in xlsx files from John Chiang / MVLGenomics Gene Panels and does the following:\n 1. Identifies hidden rows and skips these \n 2. Extracts hgvs \n\n Usage: casey_xlsx_to_hgvs.py YOURFILE.xlsx > output.tsv')
parser.add_argument('xlsx_file', help = 'John Chiang / OSU / MVL / Casey patient report xlsx file')

args = parser.parse_args()

# load workbook and read sheet names with openpyxl
wb = load_workbook(args.xlsx_file)
ws = wb[wb.get_sheet_names()[0]]
# identify header row, HGVS coding column, and zygosity column
for row in range(2, ws.max_row):
    for column in 'ABCDEFG':
        cell_name = "{}{}".format(column, row)
        if ws[cell_name].value == 'HGVSCoding':
            header_row = row
            HGVS_column = column
        if ws[cell_name].value == 'Zygosity':
            Zygosity_column = column

# get HGVS and zygosity, skipping hidden rows
hgvs_zygosity = {}
hgvs_file_name =  args.xlsx_file.split(' ')[0] + '_' + str(time.time()) + '.tmp'
hgvs_file = open(hgvs_file_name, 'w')
for row in range(header_row+1, ws.max_row):
    if ws.row_dimensions[row].visible is True:
        cell_name_hgvs = "{}{}".format(HGVS_column, row)
        cell_name_zyg = "{}{}".format(Zygosity_column, row)
        if ws[cell_name_hgvs].value is not None and ws[cell_name_hgvs].value[0:2] == 'NM':
            hgvs_zygosity[ws[cell_name_hgvs].value] = ws[cell_name_zyg].value
            hgvs_file.write(ws[cell_name_hgvs].value)
            hgvs_file.write('\n')
hgvs_file.close()

# convert to vcf with VEP
vep_call = '/Applications/ensembl-vep/./vep -i ' + hgvs_file_name + ' --port=3337 --refseq --database --force --pick --vcf --output_file STDOUT --no_stats'
vep_vcf = subprocess.check_output(vep_call, shell = True)
vep_vcf = vep_vcf.decode('utf-8')
# rm temp file
subprocess.call(['rm',hgvs_file_name])


# check if anything failed to be included (will happen if HGVS can't be parsed)
# write secondary vcf writing out the missing variants
dat_error_file_name = args.xlsx_file.split(' ')[0] + '.FAILED.dat'
dat_errorfile = open(dat_error_file_name, 'w')
for k,v in hgvs_zygosity.items():
	if k not in vep_vcf:
		dat_errorfile.write(k + '\n')

# write vcf for gemini annotation, using the zygosity to write the sample genotype
vcf_file_name = args.xlsx_file.split(' ')[0] + '.vcf'
vcf_file = open(vcf_file_name, 'w')

for line in vep_vcf.split('\n'):
    if line[0:2] == '##':
        vcf_file.write(line)
        vcf_file.write('\n')
    elif line[0:6] == '#CHROM':
        vcf_file.write('#CHROM\tPOS\tID\tREF\tALT\tQUAL\tFILTER\tINFO\tFORMAT\tSAMPLE\n')
    else:
        s_line = line.split('\t')
        if line:
            if 'het' in hgvs_zygosity[s_line[2]]:
                output =  '\t'.join(s_line[0:5]) + '\t100\tPASS\t.\tGT:GQ:DP\t' + '0/1:100:100\n'
                vcf_file.write(output)
            else:
                output = '\t'.join(s_line[0:5]) + '\t100\tPASS\t.\tGT:GQ:DP\t' + '1/1:100:100\n'
                vcf_file.write(output)