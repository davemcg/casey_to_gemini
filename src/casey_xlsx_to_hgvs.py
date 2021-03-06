#!/usr/local/Anaconda/envs/py3.4.3/bin/python

from openpyxl import load_workbook
import openpyxl
import argparse

parser = argparse.ArgumentParser(description= 'Run on biowulf2 or helix. Takes in xlsx files from John Chiang / MVLGenomics Gene Panels and extracts the following:\n    1. HGVS\n    2. HGVS marked with red (likely disease causing)\n ')
parser.add_argument('xlsx_file', help = 'John Chiang / OSU / MVL / Casey patient report xlsx file')

args = parser.parse_args()
# load workbook and read sheet names with openpyxl
wb = load_workbook(args.xlsx_file, data_only=True)
ws = wb[wb.get_sheet_names()[0]]
# identify HGVS coding column
for row in ws.iter_rows(min_row=2, max_row=20):
	for cell in row:
		if str(cell.value).upper() == 'HGVSCODING':
			header_row = cell.row
			HGVS_column = cell.column
		if str(cell.value).upper() == 'RS':
			Rs_column = cell.column
		if str(cell.value).upper() == 'CHR:CHRPOS':
			Chr_column = cell.column
		if str(cell.value).upper() == 'TIMESOBSERVEDPERPANEL':
			panel_column = cell.column
		if str(cell.value).upper() == 'TIMESOBSERVEDPERPANELGROUP':
			group_column = cell.column
		if str(cell.value).upper() == 'REF':
			ref_column = cell.column
		if str(cell.value).upper() == 'ALT':
			alt_column = cell.column

# identify last row
for row in range(header_row+1, max(ws.max_row, 5000)):
	if ws[HGVS_column + str(row)].value is None:
		last_row = row
		break

hgvs = [] # the actual hgvs
hgvs_status = [] # if red, then considered disease causing by John / MVL
rs = []
chrom_pos = []
panel = []
group = []
ref = []
alt = []

for row in range(header_row, last_row):
	hgvs_cell = "{}{}".format(HGVS_column, row)
	hgvs_cell_value = ws[hgvs_cell].value
	hgvs.append(hgvs_cell_value)
	if 'rgb=None' in str(ws[hgvs_cell].font.color):
		hgvs_status.append('Secondary')
	else:
		hgvs_status.append('Primary')
	rs.append(ws["{}{}".format(Rs_column, row)].value)
	chrom_pos.append(ws["{}{}".format(Chr_column, row)].value)
	panel.append(ws["{}{}".format(panel_column, row)].value)
	group.append(ws["{}{}".format(group_column, row)].value)
	ref.append(ws["{}{}".format(ref_column, row)].value)
	alt.append(ws["{}{}".format(alt_column, row)].value)

# print out hgvs and status (primary or secondary)
for i in range(1, len(hgvs)):
	print(hgvs[i], hgvs_status[i], rs[i], chrom_pos[i], ref[i], alt[i], panel[i], group[i])
